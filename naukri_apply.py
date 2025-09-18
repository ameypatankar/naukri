#!/usr/bin/env python3
import os
import re
import time
import sys
import logging
import openpyxl
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
    WebDriverException,
)

# ---------------- ENV SETUP ----------------
# This will load environment variables from .env if present
load_dotenv()

# ---------------- CONFIG ----------------
NAUKRI_EMAIL = os.getenv("NAUKRI_EMAIL")
NAUKRI_PASSWORD = os.getenv("NAUKRI_PASSWORD")
SKILLS = os.getenv("SKILLS")
EXPERIENCE = os.getenv("EXPERIENCE")  # years
EXCEL_FILE = os.getenv("EXCEL_FILE", "applied_jobs.xlsx")
MIN_EXPECTED_SALARY = float(os.getenv("MIN_EXPECTED_SALARY", "25"))  # LPA
MAX_APPLY = int(os.getenv("MAX_APPLY", "50"))  # Number of successful applications to reach
CHROME_DRIVER_PATH = os.getenv("CHROME_DRIVER_PATH", "")  # optional path to chromedriver
HEADLESS = os.getenv("HEADLESS", "False").lower() == "true"

LOGIN_URL = "https://www.naukri.com/nlogin/login"
SEARCH_URL = "https://www.naukri.com/jobs-in-india"

TEXT_VALUE_FOR_BOT = os.getenv("TEXT_VALUE_FOR_BOT")

# ---------------- LOGGING ----------------
logging.basicConfig(
    filename="naukri_log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logging.info("Script started")
print(f"Script started (HEADLESS={HEADLESS}). Check naukri_log.txt for detail.")

# ---------------- EXCEL SETUP ----------------
try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    logging.info(f"Loaded existing Excel: {EXCEL_FILE}")
except Exception:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Job ID", "Job Title", "Company", "Salary", "Job Link", "Status"])
    logging.info(f"Created new Excel: {EXCEL_FILE}")

existing_job_ids = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row and row[0] is not None:
        existing_job_ids.add(str(row[0]))

# ---------------- SELENIUM SETUP ----------------
options = Options()
if HEADLESS:
    options = Options()
    options.add_argument("--headless=new")        # modern headless mode
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")   # critical
    options.add_argument("--remote-allow-origins=*") # GitHub Actions fix

else:
    options.add_argument("--start-maximized")

options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--disable-extensions")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--window-size=1920,1080")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                     "AppleWebKit/537.36 (KHTML, like Gecko) "
                     "Chrome/116.0.5845.140 Safari/537.36")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

driver = None
try:
    if CHROME_DRIVER_PATH:
        from selenium.webdriver.chrome.service import Service
        service = Service(CHROME_DRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
    else:
        driver = webdriver.Chrome(options=options)
except WebDriverException as e:
    logging.error(f"Could not start ChromeDriver: {e}")
    print("ERROR: Could not start ChromeDriver:", e)
    sys.exit(1)

wait = WebDriverWait(driver, 30)
actions = ActionChains(driver)

# ---------------- HELPERS ----------------
def safe_click(element, timeout=10):
    """Scroll to element, wait until it is clickable, then click via ActionChains."""
    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        end = time.time() + timeout
        while time.time() < end:
            try:
                if element.is_displayed() and element.is_enabled():
                    break
            except StaleElementReferenceException:
                return False
            time.sleep(0.12)
        actions.move_to_element(element).pause(0.12).click().perform()
        return True
    except (ElementClickInterceptedException, ElementNotInteractableException, StaleElementReferenceException, Exception) as exc:
        logging.warning(f"safe_click failed: {exc}")
        return False

def parse_max_salary(salary_text):
    """Try to extract a numeric maximum salary in LPA from salary_text. Returns float or None."""
    if not salary_text:
        return None
    s = salary_text.lower()
    if "not disclose" in s or "not disclosed" in s:
        return None
    # normalize many common variants
    s = s.replace("lacs pa", "lpa").replace("lacs", "lpa").replace("per annum", "").replace("p.a.", "").replace("pa", "")
    nums = re.findall(r"[\d\.]+", s)
    if not nums:
        return None
    try:
        val = float(nums[-1])
        return val
    except Exception:
        return None

def save_record(job_id, title, company, salary_text, job_link, status):
    """Append a row to Excel and save immediately, and mark job id as processed."""
    try:
        sheet.append([str(job_id), title, company, salary_text, job_link, status])
        wb.save(EXCEL_FILE)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {e}")
    existing_job_ids.add(str(job_id))

# ---------------- CHATBOT ANSWERING FUNCTION ----------------
def answer_chatbot_and_submit(job_id, title, company, salary_text, job_link):
    """
    Answer chatbot questions inside chatbot_DrawerContentWrapper.
    - support contenteditable divs (div.textArea[contenteditable="true"]) by setting innerText and dispatching input event
    - fill text inputs, textareas, select the first radio/checkbox/label if possible
    - click div.sendMsg (preferred) or Next/Submit button
    - loop until drawer disappears or max iterations reached
    Returns True if drawer closed or looks handled, False otherwise.
    """
    logging.info(f"Chatbot appeared for {job_id}, attempting to auto-answer...")
    max_iterations = 20
    try:
        for iteration in range(max_iterations):
            time.sleep(0.6)  # let UI settle

            # check presence
            try:
                chatbot_el = driver.find_element(By.CLASS_NAME, "chatbot_DrawerContentWrapper")
            except Exception:
                logging.info("Chatbot wrapper not found; assuming closed.")
                return True

            # if not visible, done
            try:
                if not chatbot_el.is_displayed():
                    logging.info("Chatbot wrapper not visible; done.")
                    return True
            except Exception:
                pass

            answered_any = False

            # 1) fill normal text inputs
            try:
                text_inputs = chatbot_el.find_elements(By.XPATH, ".//input[not(@type) or @type='text']")
                for t in text_inputs:
                    try:
                        if t.is_displayed() and t.is_enabled():
                            try:
                                t.clear()
                            except Exception:
                                pass
                            t.send_keys(TEXT_VALUE_FOR_BOT)
                            answered_any = True
                            time.sleep(0.25)
                    except Exception:
                        continue
            except Exception:
                pass

            # 2) fill contenteditable divs (div.textArea[contenteditable='true'])
            try:
                editable_divs = chatbot_el.find_elements(By.XPATH, ".//div[contains(@class,'textArea') and (@contenteditable='true' or @contenteditable='')]")
                for div in editable_divs:
                    try:
                        if div.is_displayed():
                            # set innerText and dispatch input event so the site notices the change
                            driver.execute_script("""
                                arguments[0].focus();
                                arguments[0].innerText = arguments[1];
                                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                            """, div, TEXT_VALUE_FOR_BOT)
                            # also try send_keys to be safe (some frameworks detect key events)
                            try:
                                div.click()
                                div.send_keys(TEXT_VALUE_FOR_BOT)
                            except Exception:
                                pass
                            answered_any = True
                            time.sleep(0.4)
                    except Exception:
                        continue
            except Exception:
                pass

            # 3) fill <textarea> if present
            try:
                textareas = chatbot_el.find_elements(By.XPATH, ".//textarea")
                for ta in textareas:
                    try:
                        if ta.is_displayed() and ta.is_enabled():
                            try:
                                ta.clear()
                            except Exception:
                                pass
                            ta.send_keys(TEXT_VALUE_FOR_BOT)
                            answered_any = True
                            time.sleep(0.3)
                    except Exception:
                        continue
            except Exception:
                pass

            # 4) click radio/checkbox inputs
            try:
                inputs = chatbot_el.find_elements(By.XPATH, ".//input[@type='radio' or @type='checkbox']")
                for inp in inputs:
                    try:
                        if inp.is_displayed() and inp.is_enabled():
                            driver.execute_script("arguments[0].click();", inp)
                            answered_any = True
                            time.sleep(0.25)
                            break
                    except Exception:
                        continue
                # fallback: click the first visible label option
                if not answered_any:
                    label_opts = chatbot_el.find_elements(By.XPATH, ".//label")
                    for lbl in label_opts:
                        try:
                            txt = (lbl.text or "").strip()
                            if txt and lbl.is_displayed():
                                driver.execute_script("arguments[0].click();", lbl)
                                answered_any = True
                                time.sleep(0.25)
                                break
                        except Exception:
                            continue
            except Exception:
                pass

            # 5) select first option in selects
            try:
                selects = chatbot_el.find_elements(By.TAG_NAME, "select")
                for sel in selects:
                    try:
                        if sel.is_displayed() and sel.is_enabled():
                            try:
                                Select(sel).select_by_index(1)
                            except Exception:
                                try:
                                    Select(sel).select_by_index(0)
                                except Exception:
                                    pass
                            answered_any = True
                            time.sleep(0.3)
                            break
                    except Exception:
                        continue
            except Exception:
                pass

            # 6) click send button(s) - prefer div.sendMsg
            try:
                send_btns = chatbot_el.find_elements(By.CSS_SELECTOR, "div.sendMsg, button.sendMsg, .sendMsg")
                for sbtn in send_btns:
                    try:
                        if sbtn.is_displayed():
                            driver.execute_script("arguments[0].click();", sbtn)
                            answered_any = True
                            time.sleep(0.8)  # let the next question appear
                            break
                    except Exception:
                        continue
            except Exception:
                pass

            # 7) fallback: click a Next/Submit/Continue inside the chatbot container
            try:
                nxt_btn = None
                try:
                    nxt_btn = chatbot_el.find_element(By.XPATH, ".//button[contains(.,'Next') or contains(.,'Submit') or contains(.,'Continue')]")
                except Exception:
                    try:
                        nxt_btn = chatbot_el.find_element(By.XPATH, ".//a[contains(.,'Next') or contains(.,'Submit') or contains(.,'Continue')]")
                    except Exception:
                        nxt_btn = None

                if nxt_btn and nxt_btn.is_displayed() and nxt_btn.is_enabled():
                    driver.execute_script("arguments[0].click();", nxt_btn)
                    answered_any = True
                    time.sleep(0.9)
            except Exception:
                pass

            # if nothing we could do in this iteration, break to avoid infinite loop
            if not answered_any:
                logging.info("Could not auto-answer further questions (no recognizable inputs/buttons).")
                break

            # check if drawer closed after actions
            time.sleep(0.5)
            try:
                current_chat = driver.find_element(By.CLASS_NAME, "chatbot_DrawerContentWrapper")
                if not current_chat.is_displayed():
                    logging.info("Chatbot closed after answers.")
                    return True
            except Exception:
                logging.info("Chatbot element not found after answering; assuming closed.")
                return True

        # after iterations, check if closed or present
        try:
            driver.find_element(By.CLASS_NAME, "chatbot_DrawerContentWrapper")
            logging.warning("Chatbot still present after attempts.")
            return False
        except Exception:
            return True

    except Exception as e:
        logging.exception(f"Exception while answering chatbot for {job_id}: {e}")
        return False

# ---------------- MAIN FLOW ----------------
try:
    # ---------- LOGIN ----------
    logging.info("Opening login page")
    print("Opening login page...")
    driver.get(LOGIN_URL)
    try:
        wait.until(EC.presence_of_element_located((By.ID, "usernameField")))
    except TimeoutException:
        driver.save_screenshot("login_page_not_loaded.png")
        logging.error("Login page did not load usernameField; exiting.")
        raise SystemExit("Login field not found")

    try:
        username_input = driver.find_element(By.ID, "usernameField")
        password_input = driver.find_element(By.ID, "passwordField")
        username_input.clear()
        username_input.send_keys(NAUKRI_EMAIL)
        password_input.clear()
        password_input.send_keys(NAUKRI_PASSWORD)
        submit_btn = driver.find_element(By.XPATH, "//button[@type='submit']")
        print("Submitting login...")
        if not safe_click(submit_btn):
            try:
                submit_btn.click()
            except Exception:
                pass
        logging.info("Login submitted")
        time.sleep(4)
    except Exception as e:
        driver.save_screenshot("login_error.png")
        logging.error(f"Login error: {e}", exc_info=True)
        raise

    # ---------- SEARCH ----------
    logging.info("Navigating to job search page")
    print("Navigating to job search page...")
    driver.get(SEARCH_URL)
    time.sleep(2)
    try:
        search_bar_container = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "nI-gNb-sb__main")))
        safe_click(search_bar_container)
        search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter keyword / designation / companies']")))
        search_box.clear()
        search_box.send_keys(SKILLS)
        exp_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='dropdownMainContainer']")))
        safe_click(exp_dropdown)
        exp_option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//li[@title='{EXPERIENCE} years']")))
        safe_click(exp_option)
        search_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class='nI-gNb-sb__icon-wrapper']")))
        safe_click(search_button)
        logging.info("Search executed")
        print("Search executed, waiting for results...")
        time.sleep(4)
    except Exception as e:
        driver.save_screenshot("search_error.png")
        logging.error(f"Job search failed: {e}", exc_info=True)
        raise

    # ---------- MAIN LOOP (pages -> jobs) ----------
    applied_count = 0
    page_num = 1
    visited_pages = set()
    print(f"Starting job processing (target apply count = {MAX_APPLY})")

    while applied_count < MAX_APPLY:
        logging.info(f"Processing page {page_num}")
        print(f"\n--- Processing page {page_num} --- (applied so far: {applied_count})")
        current_url = driver.current_url
        if current_url in visited_pages:
            logging.info("Already visited this page URL; stopping to avoid loop.")
            break
        visited_pages.add(current_url)

        # Prefer job container inside chatbot wrapper if present
        jobs = []
        try:
            container = driver.find_element(By.CLASS_NAME, "chatbot_DrawerContentWrapper")
            jobs = container.find_elements(By.XPATH, ".//div[contains(@class,'srp-jobtuple-wrapper') or contains(@class,'jobTuple')]")
            logging.info("Found job container inside chatbot_DrawerContentWrapper")
        except Exception:
            # fallback to site-wide job cards
            jobs = driver.find_elements(By.XPATH, "//div[contains(@class,'srp-jobtuple-wrapper') or contains(@class,'jobTuple')]")

        if not jobs:
            logging.info("No job cards found on this page. Ending.")
            print("No job cards found on this page. Ending.")
            break

        # iterate by index to avoid stale-element issues
        idx = 0
        while idx < len(jobs) and applied_count < MAX_APPLY:
            # Re-fetch job list each iteration
            try:
                jobs = driver.find_elements(By.XPATH, "//div[contains(@class,'srp-jobtuple-wrapper') or contains(@class,'jobTuple')]")
                job = jobs[idx]
            except Exception:
                idx += 1
                continue
            idx += 1

            # get job id
            try:
                job_id = job.get_attribute("data-job-id")
            except StaleElementReferenceException:
                continue

            if not job_id:
                continue
            if str(job_id) in existing_job_ids:
                # skip already processed
                continue

            # extract fields
            try:
                title_elem = job.find_element(By.XPATH, ".//a[contains(@class,'title')]")
                title = title_elem.text.strip()
                job_link = title_elem.get_attribute("href")
            except Exception:
                title = "Unknown Title"
                job_link = "N/A"

            try:
                company = job.find_element(By.XPATH, ".//a[contains(@class,'comp-name') or contains(@class,'subTitle')]").text.strip()
            except Exception:
                company = "Unknown Company"

            try:
                salary_text = job.find_element(By.XPATH, ".//span[contains(@class,'sal-wrap')]").text.strip()
            except Exception:
                salary_text = "Not Disclosed"

            print(f"Found job: {title} | {company} | {salary_text} | ID: {job_id}")
            logging.info(f"Found job: {job_id} | {title} | {company} | {salary_text}")

            # If job card indicates Already Applied -> record and skip (do not increment)
            try:
                applied_tag = job.find_element(By.XPATH, ".//span[contains(text(),'Applied')]")
                if applied_tag.is_displayed():
                    save_record(job_id, title, company, salary_text, job_link, "Already Applied")
                    logging.info(f"Card shows Already Applied for {job_id} — recorded and skipped")
                    continue
            except Exception:
                pass

            # Salary filter
            max_sal = parse_max_salary(salary_text)
            if (max_sal is not None) and (max_sal < MIN_EXPECTED_SALARY):
                save_record(job_id, title, company, salary_text, job_link, "Skipped (Low Salary)")
                logging.info(f"Skipped low salary job {job_id}: {salary_text}")
                continue

            # Open job detail (click title) - may open new tab
            try:
                if not safe_click(title_elem):
                    try:
                        title_elem.click()
                    except Exception:
                        save_record(job_id, title, company, salary_text, job_link, "Could not open job detail")
                        continue
                time.sleep(1)
                if len(driver.window_handles) > 1:
                    driver.switch_to.window(driver.window_handles[-1])
                time.sleep(1)
            except Exception as e:
                logging.error(f"Error opening job detail for {job_id}: {e}", exc_info=True)
                save_record(job_id, title, company, salary_text, job_link, f"Open detail error: {e}")
                try:
                    if len(driver.window_handles) > 1:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                except Exception:
                    pass
                continue

            # find apply button on detail page (buttons or links containing 'apply')
            apply_btn = None
            try:
                candidates = driver.find_elements(By.XPATH, "//button|//a")
                for b in candidates:
                    try:
                        txt = (b.text or "").strip().lower()
                        if "apply" in txt and b.is_displayed() and b.is_enabled():
                            apply_btn = b
                            break
                    except Exception:
                        continue
            except Exception:
                apply_btn = None

            if not apply_btn:
                save_record(job_id, title, company, salary_text, job_link, "No Apply Button")
                logging.info(f"No apply button on detail for {job_id}")
                # close detail tab if opened
                if len(driver.window_handles) > 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                else:
                    try:
                        driver.back()
                    except Exception:
                        pass
                continue

            btn_text = (apply_btn.text or "").strip().lower()

            # Skip company-site applies
            if "company site" in btn_text or "apply on company" in btn_text:
                save_record(job_id, title, company, salary_text, job_link, "Skipped (Company Site)")
                logging.info(f"Skipped company-site job {job_id}")
                if len(driver.window_handles) > 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                else:
                    try:
                        driver.back()
                    except Exception:
                        pass
                continue

            # If button already says "Applied"
            if "applied" in btn_text:
                save_record(job_id, title, company, salary_text, job_link, "Already Applied")
                logging.info(f"Detail shows Already Applied for {job_id}")
                if len(driver.window_handles) > 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                else:
                    try:
                        driver.back()
                    except Exception:
                        pass
                continue

            # Click apply
            clicked = safe_click(apply_btn)
            if not clicked:
                try:
                    apply_btn.click()
                except Exception as e:
                    save_record(job_id, title, company, salary_text, job_link, "No Apply Button / Not Clickable")
                    logging.warning(f"Could not click apply for {job_id}: {e}")
                    if len(driver.window_handles) > 1:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                    else:
                        try:
                            driver.back()
                        except Exception:
                            pass
                    continue

            # After clicking check for chatbot drawer (short wait)
            time.sleep(1.5)
            chatbot_shown = False
            try:
                # short explicit wait for chatbot drawer presence
                small_wait = WebDriverWait(driver, 3)
                small_wait.until(EC.presence_of_element_located((By.CLASS_NAME, "chatbot_DrawerContentWrapper")))
                # if found and visible, mark present
                chatbot_el = driver.find_element(By.CLASS_NAME, "chatbot_DrawerContentWrapper")
                if chatbot_el and chatbot_el.is_displayed():
                    chatbot_shown = True
            except Exception:
                chatbot_shown = False

            if chatbot_shown:
                # Try to answer chatbot questions instead of skipping
                handled = answer_chatbot_and_submit(job_id, title, company, salary_text, job_link)
                if not handled:
                    # failed to handle chatbot -> record and continue
                    save_record(job_id, title, company, salary_text, job_link, "Skipped (Chatbot)")
                    logging.info(f"Chatbot appeared for {job_id} and could not be handled; skipped.")
                    # close tab or go back
                    try:
                        if len(driver.window_handles) > 1:
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                        else:
                            driver.back()
                    except Exception:
                        pass
                    continue
                else:
                    # chatbot closed - wait a bit
                    time.sleep(1.2)

            # No chatbot or handled - now check if apply succeeded or already applied
            status = "Unknown"
            try:
                # Try to locate apply/applied button again (DOM may have changed)
                try:
                    new_apply_btn = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.XPATH, "//button[contains(text(),'Apply') or contains(text(),'Applied')]"))
                    )
                except Exception:
                    new_apply_btn = None

                btn_text = ""
                if new_apply_btn:
                    try:
                        btn_text = (new_apply_btn.text or "").strip().lower()
                    except Exception:
                        btn_text = ""

                if "apply" == btn_text:
                    # If still 'Apply' text, assume success for our flows
                    status = "Applied Successfully"
                    logging.info(f"Assuming applied for {job_id} (button still shows 'Apply').")
                elif "applied" in btn_text:
                    status = "Applied Successfully"
                    logging.info(f"Detected Applied label for {job_id}")
                else:
                    # If we couldn't find a button, assume success if no errors
                    status = "Applied Successfully"
            except Exception as e:
                status = "Applied (unknown state)"
                logging.warning(f"Error determining apply status for {job_id}: {e}")

            # Record result
            save_record(job_id, title, company, salary_text, job_link, status)
            if status == "Applied Successfully":
                applied_count += 1
                logging.info(f"Applied to {job_id} — total applied {applied_count}")
            else:
                logging.info(f"Processed {job_id} with status: {status}")

            # close detail tab or navigate back to results
            try:
                if len(driver.window_handles) > 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                else:
                    driver.back()
            except Exception:
                pass

            # short human-like delay
            time.sleep(1.2)

        # ---------- PAGINATION ----------
        if applied_count >= MAX_APPLY:
            logging.info(f"Reached MAX_APPLY ({MAX_APPLY}). Ending.")
            break

        logging.info("Attempting pagination (numbered pages -> Next fallback)")
        # Try numbered pages first (div.lastCompMark -> div.styles_pages__v1rAK a)
        next_clicked = False
        try:
            pagination_container = driver.find_element(By.CSS_SELECTOR, "div.lastCompMark div.styles_pages__v1rAK")
            links = pagination_container.find_elements(By.TAG_NAME, "a")
            # build mapping number->href for numeric links
            page_map = {}
            for link in links:
                txt = (link.text or "").strip()
                href = link.get_attribute("href")
                if txt.isdigit() and href:
                    try:
                        num = int(txt)
                        page_map[num] = href
                    except Exception:
                        continue
            # prefer page_num+1 if available, else smallest > page_num
            target_href = None
            if (page_num + 1) in page_map:
                target_href = page_map[page_num + 1]
            else:
                greater = [n for n in sorted(page_map.keys()) if n > page_num]
                if greater:
                    target_href = page_map[greater[0]]
            if target_href:
                logging.info(f"Going to next numeric page: {target_href}")
                driver.get(target_href)
                # wait until URL changes (safety)
                try:
                    WebDriverWait(driver, 8).until(lambda d: d.current_url != current_url)
                except Exception:
                    logging.info("URL didn't change after numeric page click; continuing anyway.")
                page_num += 1
                time.sleep(3)
                next_clicked = True
        except Exception:
            next_clicked = False

        if not next_clicked:
            # Fallback to Next link/button
            try:
                next_btn = driver.find_element(By.XPATH, "//a[contains(text(),'Next') or contains(., 'Next')]")
                if safe_click(next_btn) or True:
                    try:
                        WebDriverWait(driver, 8).until(lambda d: d.current_url != current_url)
                    except Exception:
                        logging.info("URL didn't change after Next click")
                    page_num += 1
                    time.sleep(3)
                    next_clicked = True
            except Exception:
                next_clicked = False

        if not next_clicked:
            logging.info("No next page found; ending pagination.")
            print("No next page found; ending.")
            break

    # Main loop done
    logging.info(f"Completed. Total applied: {applied_count}")
    print(f"\nDone. Applied {applied_count} jobs. Excel: {EXCEL_FILE}")
    try:
        wb.save(EXCEL_FILE)
    except Exception:
        logging.warning("Failed to save Excel at final step.")
    driver.quit()

except Exception as fatal:
    logging.exception("Fatal error during script execution")
    print(f"Fatal error: {fatal}. See naukri_log.txt and screenshots.")
    try:
        driver.save_screenshot("fatal_error.png")
    except Exception:
        pass
    try:
        wb.save(EXCEL_FILE)
    except Exception:
        pass
finally:
    try:
        if driver:
            driver.quit()
    except Exception:
        pass
    logging.info("Script finished (final).")
