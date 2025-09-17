from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl

# ---- CONFIG ----
NAUKRI_EMAIL = "ameypatankar333@gmail.com"
NAUKRI_PASSWORD = "9773051915"
SKILLS = "Java and Spring Boot And React"
EXPERIENCE = "11"  # in years
CHROME_DRIVER_PATH = r"C:\Users\Amey\Downloads\139.0.7258.66 chromedriver-win64\chromedriver-win64\chromedriver.exe"
EXCEL_FILE = "applied_jobs.xlsx"
MIN_EXPECTED_SALARY = 25  # LPA (Minimum salary expectation)
MAX_APPLY = 10  # Number of new jobs to process

# ---- START ----
service = Service(CHROME_DRIVER_PATH)
driver = webdriver.Chrome(service=service)
driver.maximize_window()
wait = WebDriverWait(driver, 20)

# ---- CREATE / LOAD EXCEL ----
try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Job ID", "Job Title", "Company", "Salary", "Job Link", "Status"])  # Header row

# Load existing job IDs to avoid duplicates
existing_job_ids = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    existing_job_ids.add(str(row[0]))

try:
    # 1️⃣ Open Naukri Login Page
    driver.get("https://www.naukri.com/nlogin/login")
    time.sleep(3)

    # 2️⃣ Login
    driver.find_element(By.ID, "usernameField").send_keys(NAUKRI_EMAIL)
    driver.find_element(By.ID, "passwordField").send_keys(NAUKRI_PASSWORD)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(5)

    # 3️⃣ Go to Job Search
    driver.get("https://www.naukri.com/jobs-in-india")

    # --- STEP 2: Click Search Bar Container ---
    search_bar_container = wait.until(
        EC.element_to_be_clickable((By.CLASS_NAME, "nI-gNb-sb__main"))
    )
    search_bar_container.click()

    # --- STEP 3: Enter Skills ---
    search_box = wait.until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter keyword / designation / companies']"))
    )
    search_box.clear()
    search_box.send_keys(SKILLS)

    # 4️⃣ Click Experience Dropdown
    exp_dropdown = wait.until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='dropdownMainContainer']"))
    )
    exp_dropdown.click()

    exp_dropdown_list = wait.until(
        EC.element_to_be_clickable((By.XPATH, f"//li[@title='{EXPERIENCE} years']"))
    )
    driver.execute_script("arguments[0].click();", exp_dropdown_list)

    # 5️⃣ Click Search Button
    search_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//button[@class='nI-gNb-sb__icon-wrapper']"))
    )
    search_button.click()
    time.sleep(5)
    print("📋 Fetching job list...")

    # ✅ Loop through jobs until MAX_APPLY new jobs processed
    applied_count = 0
    while applied_count < MAX_APPLY:
        jobs = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//div[@class='srp-jobtuple-wrapper']"))
        )
        if not jobs:
            print("No jobs found.")
            break

        for job in jobs:
            # Get Job ID
            job_id = job.get_attribute("data-job-id")
            if job_id in existing_job_ids:
                print(f"⏭️ Skipping Job ID {job_id} (already processed)")
                continue

            # Get job details
            try:
                title_elem = job.find_element(By.XPATH, ".//a[contains(@class,'title')]")
                title = title_elem.text.strip()
                job_link = title_elem.get_attribute("href")
            except:
                title = "Unknown Title"
                job_link = "N/A"

            try:
                company = job.find_element(By.XPATH, ".//a[contains(@class,'comp-name')]").text.strip()
            except:
                company = "Unknown Company"

            try:
                salary_elem = job.find_element(By.XPATH, ".//span[contains(@class,'sal-wrap')]")
                salary_text = salary_elem.text.strip()
            except:
                salary_text = "Not Disclosed"

            print(f"\n🔎 Job: {title} | {company} | {salary_text} | ID: {job_id}")

            # Check if already applied in job card
            try:
                status_tag = job.find_element(By.XPATH, ".//span[contains(text(),'Applied')]")
                if status_tag.is_displayed():
                    status = "Already Applied"
                    print(f"⏭️ Job ID {job_id} already applied")
                    sheet.append([job_id, title, company, salary_text, job_link, status])
                    wb.save(EXCEL_FILE)
                    existing_job_ids.add(job_id)
                    continue
            except:
                pass

            # Salary filtering
            apply_job = True
            if "Not Disclosed" not in salary_text:
                try:
                    salary_parts = salary_text.replace("Lacs PA", "").strip().split("-")
                    min_salary = float(salary_parts[0])
                    max_salary = float(salary_parts[1]) if len(salary_parts) > 1 else min_salary
                    if max_salary < MIN_EXPECTED_SALARY:
                        apply_job = False
                except:
                    apply_job = True

            if not apply_job:
                status = "Skipped (Low Salary)"
                sheet.append([job_id, title, company, salary_text, job_link, status])
                wb.save(EXCEL_FILE)
                existing_job_ids.add(job_id)
                continue

            # Click job to apply
            job.click()
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(2)

            try:
                apply_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Apply') or contains(text(),'Applied')]"))
                )
                apply_btn.click()
                time.sleep(2)

                # Check for chatbot drawer
                try:
                    chatbot = driver.find_element(By.CLASS_NAME, "chatbot_DrawerContentWrapper")
                    if chatbot.is_displayed():
                        status = "Skipped (Chatbot)"
                        print(f"🤖 Job ID {job_id} skipped due to chatbot drawer.")
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        sheet.append([job_id, title, company, salary_text, job_link, status])
                        wb.save(EXCEL_FILE)
                        existing_job_ids.add(job_id)
                        continue
                except:
                    # No chatbot, check button text
                    btn_text = apply_btn.text.strip().lower()
                    if btn_text == "apply":
                        status = "Applied Successfully"
                        print(f"✅ Applied Job ID {job_id}")
                    else:
                        status = "Already Applied"
                        print(f"⏭️ Already applied Job ID {job_id}")

            except:
                status = "No Apply Button"
                print(f"⚠️ No apply button found for Job ID {job_id}")

            # Save result
            sheet.append([job_id, title, company, salary_text, job_link, status])
            wb.save(EXCEL_FILE)
            existing_job_ids.add(job_id)

            # Close tab & switch back
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            applied_count += 1  # ✅ Increase count only for actually processed job

            if applied_count >= MAX_APPLY:
                break

except Exception as e:
    print("❌ Error:", e)
finally:
    try:
        wb.save(EXCEL_FILE)
    except:
        print("⚠️ Could not save file (might be open). Close it and rerun.")
    driver.quit()
